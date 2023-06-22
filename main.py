import tkinter as tk
from tkinter import ttk
import openpyxl
import datetime

# ///////////////////////////////////////////////////////////////////////////////////////////////////////////////

# Variables
data_discount = ''
workbook_num = 0


# LOAD DATA FUNCTION
def load_data():
    path = "prices.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    list_values = list(sheet.values)
    for col_name in list_values[0]:
        treeview.heading(col_name, text=col_name)

    for value_tuple in list_values[1:]:
        treeview.insert('', tk.END, values=value_tuple)


# INSERT ROW FUNCTION
def insert_row():
    only_numbers_label.grid_forget()
    # GET VALUES
    global data_discount
    date = datetime.date.today()
    try:
        price = int(price_entry.get())
        discount = data_discount
        event = event_entry.get()
        if event == "Add a comment":
            event = 'No comment'
        else:
            pass

        # INSERT ROW INTO EXCEL SHEET
        path = "prices.xlsx"
        workbook = openpyxl.load_workbook(path)
        workbook.active = workbook[workbook_num]
        sheet = workbook.active
        row_values = [date, price, discount, event]
        sheet.append(row_values)
        workbook.save(path)

        # SHOW ON TREEVIEW
        treeview.insert('', tk.END, values=row_values)

        # CLEAR THE VALUES
        price_entry.delete(0, 'end')
        price_entry.insert(0, 'PRICE')
        discount_menu.config(text=discount_list[0])
        event_entry.delete(0, 'end')
        event_entry.insert(0, 'Add a comment')
    except ValueError:
        only_numbers_label.grid(row=1, column=1, padx=10, pady=(10, 0))
        pass


# Callback function
def callback(selection):
    global data_discount
    data_discount = selection


# DELETE ROW FUNCTION
def delete_row():
    row_numb = int(spin_box.get()) + 1
    if row_numb == 1:
        pass
    else:
        path = "prices.xlsx"
        workbook = openpyxl.load_workbook(path)
        workbook.active = workbook[workbook_num]
        sheet = workbook.active
        sheet.delete_rows(row_numb)
        workbook.save(path)

        # update treeview
        treeview.delete(*treeview.get_children())
        load_data()

        # clear values
        spin_box.delete(0, 'end')
        spin_box.insert(0, "Select Row")


# SELECT SHEET
def select_sheet(selection):
    # Update Information
    treeview.delete(*treeview.get_children())

    path = "prices.xlsx"
    workbook = openpyxl.load_workbook(path)
    workbook.active = workbook[selection]
    global workbook_num
    workbook_num = selection
    sheet = workbook.active

    list_values = list(sheet.values)
    for col_name in list_values[0]:
        treeview.heading(col_name, text=col_name)

    for value_tuple in list_values[1:]:
        treeview.insert('', tk.END, values=value_tuple)


def get_sheets_names():
    menu = ['Select a product']
    path = "prices.xlsx"
    workbook = openpyxl.load_workbook(path)
    for sheet in workbook.sheetnames:
        menu.append(sheet)
    return menu


def create_product():
    # Get Product_name
    ws_name = add_product_entry.get()
    # Delete current values on treeview
    treeview.delete(*treeview.get_children())

    # Add new Sheet to the file
    path = "prices.xlsx"
    workbook = openpyxl.load_workbook(path)
    workbook.create_sheet(ws_name)
    workbook.save(path)
    workbook.active = workbook[ws_name]
    sheet = workbook.active
    row_values = ['Date', 'Price', 'Discount', 'Comment']
    sheet.append(row_values)
    workbook.save(path)

    # Reset the values on product entry
    add_product_entry.delete(0, 'end')
    add_product_entry.insert(0, 'New Product Name')


# ///////////////////////////////////////////////////////////////////////////////////////////////////////////////

# MAIN
root = tk.Tk()
root.title('Price tracker')

# Import the tcl file
root.tk.call('source', 'forest-dark.tcl')

# Set the theme with the theme_use method
ttk.Style().theme_use('forest-dark')


# A (ttk) frame
frame = ttk.Frame(root)
frame.grid(row=0, column=0)

# ///////////////////////////////////////////////////////////////////////////////////////////////////////////////

# Select sheet list

# Select sheet frame

select_item_frame = ttk.LabelFrame(frame, text='Choose Sheet')
select_item_frame.grid(row=0, column=0, pady=(20, 0), padx=10, sticky='ew')

# Menu List

menu_list = get_sheets_names()

# Variables
e = tk.StringVar(value=menu_list[0])

select_item_list = ttk.OptionMenu(select_item_frame, e, *menu_list, command=select_sheet)
select_item_list.grid(row=0, column=0, pady=10, padx=10)

# ADD NEW PRODUCT:
add_product_label = ttk.Label(select_item_frame, text='Add New Product:')
add_product_label.grid(row=1, column=0, pady=(0, 10), padx=10)

add_product_entry = ttk.Entry(select_item_frame)
add_product_entry.insert(0, "New Product Name")
add_product_entry.bind("<FocusIn>", lambda a: add_product_entry.delete('0', 'end'))
add_product_entry.grid(row=1, column=1, pady=(0, 10), padx=10)

add_product_btn = ttk.Button(select_item_frame, text="Add New Item", command=create_product)
add_product_btn.grid(row=1, column=2, pady=(0, 10), padx=10)

# ///////////////////////////////////////////////////////////////////////////////////////////////////////////////

# Ttk label frame
widgets_frame = ttk.LabelFrame(frame, text="Insert Row")
widgets_frame.grid(row=1, column=0, pady=10, padx=10, sticky='ew')


# Entry Widget PRICE
price_entry = ttk.Entry(widgets_frame)
price_entry.insert(0, "PRICE")
price_entry.bind("<FocusIn>", lambda a: price_entry.delete('0', 'end'))
price_entry.grid(row=1, column=0, sticky='ew', pady=(10, 0), padx=10)


# ///////////////////////////////////////////////////////////////////////////////////////////////////////////////

# Select Discount list

# Discount list
discount_list = ['Does it have discount?', 'Discount', 'No discount']

# Variables
f = tk.StringVar(value=menu_list[0])

discount_menu = ttk.OptionMenu(widgets_frame, f, *discount_list, command=callback)
discount_menu.grid(row=2, column=0, sticky='ew', pady=(20, 0), padx=10)

# Event entry

# Entry Widget
event_entry = ttk.Entry(widgets_frame)
event_entry.insert(0, "Add a comment")
event_entry.bind("<FocusIn>", lambda a: event_entry.delete('0', 'end'))
event_entry.grid(row=3, column=0, sticky='ew', pady=(20, 0), padx=10)

# BUTTON
button_insert_information = ttk.Button(widgets_frame, text="Insert", command=insert_row)
button_insert_information.grid(row=4, column=0, sticky='ew', pady=(20, 20), padx=10)

# Only numbers label

only_numbers_label = ttk.Label(widgets_frame, text="You can only type numbers")


# ///////////////////////////////////////////////////////////////////////////////////////////////////////////////

# Creating the treeFrame and treeview

treeFrame = ttk.Frame(frame)
treeFrame.grid(row=0, column=1, pady=10, padx=10, rowspan=4)

# treeScroll
treeScroll = ttk.Scrollbar(treeFrame)
treeScroll.pack(side="right", fill="y")

# Treeview
cols = ("Date", "Price", "Discount", "Comment")
treeview = ttk.Treeview(treeFrame, show="headings", columns=cols, height=13,
                        yscrollcommand=treeScroll.set)
treeview.column('Date', width=100)
treeview.column('Price', width=100)
treeview.column('Discount', width=100)
treeview.column('Comment', width=400)
treeview.pack()

# Configurate the scroll to affect the treeview
treeScroll.config(command=treeview.yview)

load_data()

# ///////////////////////////////////////////////////////////////////////////////////////////////////////////////

# DELETE ROW FRAME
delete_row_frame = ttk.LabelFrame(frame, text="Delete Row")
delete_row_frame.grid(row=2, column=0, pady=(0, 10), padx=10, sticky='ew')

# SELECT ROW OPTION

spin_box = ttk.Spinbox(delete_row_frame, from_=1, to=1000)
spin_box.insert(0, "Select Row")
spin_box.bind("<FocusIn>", lambda a: spin_box.delete('0', 'end'))
spin_box.grid(row=0, column=0, pady=10, padx=10)
delete_row_button = ttk.Button(delete_row_frame, text="Delete Row", command=delete_row)
delete_row_button.grid(row=0, column=1, pady=10, padx=(0, 10))

root.mainloop()
